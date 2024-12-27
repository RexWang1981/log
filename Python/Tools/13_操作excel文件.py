# Create a workbook

from openpyxl import Workbook
from openpyxl import load_workbook
# 建立空的工作表  #（ Workbook class）
wb = Workbook()
# 建立空的work sheet （Workbook.active property）
ws = wb.active

# You can create new worksheets using the Workbook.create_sheet() method:

# insert at the end (default)， or
ws1 = wb.create_sheet("2nd_sheet_end")

# insert at first position, or
ws2 = wb.create_sheet("3rd_sheet_front", 0)

# insert at the penultimate position, or
ws3 = wb.create_sheet("4th_sheet_In_between", -1)

# 更改work sheet的title
ws.title = "1st_sheet"
ws1.title = "2nd_sheet_end-2"
ws2.title = "3rd_sheet_front-2"
ws3.title = "4th_sheet_In_between-2"

# You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute
print(wb.sheetnames)

# Once you gave a worksheet a name, you can get it as a key of the workbook
print(wb["1st_sheet"])

# You can loop through worksheets
for sheet in wb:
    print(sheet.title)

# You can create copies of worksheets within a single workbook:
source = wb.active
target = wb.copy_worksheet(source)

"""
Note Only cells (including values, styles, hyperlinks and comments) and certain worksheet attributes (including dimensions, format and properties) are copied. All other workbook / worksheet attributes are not copied - e.g. Images, Charts.
You also cannot copy worksheets between workbooks. You cannot copy a worksheet if the workbook is open in read-only or write-only mode.
"""
# -------------------------------------------------------------
# Accessing one cell

# 给worksheet ws1的A4单元格赋值为4
ws1["A4"] = 4
# 读取ws1的A4单元格的值
c = ws1['A4']

print(c)

# 也可以用cell（）方法查找单元格数值，如下
d = ws.cell(row=4, column=2, value=10)
print(d)

# -------------------------------------------------------------
# Accessing many cells
cell_range = ws['A1':'C2']

# Ranges of rows or columns can be obtained similarly:
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

# You can also use the Worksheet.iter_rows() method:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
   for cell in row:
       print(cell)

# Likewise the Worksheet.iter_cols() method will return columns: Note For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.
#For col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
#    for cell in col:
 #       print(cell)

#遍历别个行列。
ws = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
# or
# Worksheet.columns property:
tuple(ws.columns)


# you can use the Worksheet.values property.获得表格的值
for row in ws.values:
   for value in row:
     print(value)

# Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
  print(row)

# Data storage, Once we have a Cell, we can assign it a value:
c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)

# Saving to a file
# The simplest and safest way to save a workbook is by using the Workbook.save() method of the Workbook object:
wb = Workbook()
# Warning This operation will overwrite existing files without warning.
wb.save('balances.xlsx')

# If required, you can specify the attribute wb.template=True, to save a workbook as a template:
# Note The filename extension is not forced to be xlsx or xlsm, although you might have some trouble opening it directly with another application if you don’t use an official extension.
# As OOXML files are basically ZIP files, you can also open it with your favourite ZIP archive manager.

wb = load_workbook('Prompt.xlsx')
wb.template = True
wb.save('Prompt_template.xltx')

# Loading from a file
# You can use the openpyxl.load_workbook() to open an existing workbook:

wb = load_workbook(filename = 'Prompt.xlsx')
sheet_ranges = wb['Tabelle1']
print(sheet_ranges['D18'].value)
'''
Note There are several flags that can be used in load_workbook.
data_only controls whether cells with formulae have either the
formula (default) or the value stored the last time Excel read the sheet.

keep_vba controls whether any Visual Basic elements are preserved or
not (default). If they are preserved they are still not editable.

read-only opens workbooks in a read-only mode. This uses much less
memory and is faster but not all features are available (charts, images, etc.)

rich_text controls whether any rich-text formatting in cells is
preserved. The default is False.

keep_links controls whether data cached from external workbooks is
preserved.
'''
